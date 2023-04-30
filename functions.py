
import time
import sys

import openpyxl
import pandas as pd

import win32com.client


def refresh_tracker():

    try:
        # Start an instance of Excel
        xlapp = win32com.client.DispatchEx("Excel.Application")
        # Open the workbook in said instance of Excel
        wb = xlapp.workbooks.open(r'C:\Users\klecznar\Desktop\SENDER\Copy of Collins.xlsx')
        # Show Excel
        xlapp.Visible = True
        # Refresh all data connections.
        wb.RefreshAll()
        xlapp.DisplayAlerts = False
        time.sleep(10)
        wb.Save()
        time.sleep(10)
        wb.Close(False)
        xlapp.DisplayAlerts = True
        # Quit
        xlapp.Quit()
    except Exception as e:
        print("SOMETHING WENT WRONG WHILE REFRESHING QUERY..." + e)


def analyze_rows():

    try:
        excel = openpyxl.load_workbook('Copy of Collins.xlsx')

        # create lists to hold info
        NI_No = []
        PN_list = []
        supplier_email = []
        date_list = []

        for i in range(1048575):  # 1,048,576 is max number of rows in excel
            # get NI FAIR number
            NI_FAIR_No = excel['Query2'].cell(row=11 + i, column=2).value
            # get part number starting from row 2
            part_number = excel['Query2'].cell(row=11 + i, column=3).value
            # get supplier e-mail
            supplier_email_item = excel['Query2'].cell(row=11 + i, column=19).value
            # break out of loop if cell is empty
            if part_number is None or part_number == '':
                break
            # check status of FAIR (must be 'awaiting')
            FAIR_status = excel['Query2'].cell(row=11 + i, column=15).value
            if FAIR_status == 'awaiting':
                # check the date of status change
                date = excel['Query2'].cell(row=11 + i, column=13).value
                # get number of days that passed since today
                No_days = excel['Query2'].cell(row=11 + i, column=14).value
                if No_days >= 5 or No_days == 0:
                    NI_No.append(NI_FAIR_No)
                    PN_list.append(part_number)
                    supplier_email.append(supplier_email_item)
                    date_list.append(date)


        if not PN_list:
            print("Nothing to escalate...")
            sys.exit()
        else:
            print("Sending mail...")

    except Exception as e:
        print("ERROR: " + e)

    return NI_No, PN_list, supplier_email, date_list


def send_mail():

    # access lists from previous function
    NI_No, PN_list, supplier_email, date_list = analyze_rows()

    # create an instance of Outlook
    olapp = win32com.client.DispatchEx("Outlook.Application")


    for item, PN, email, date in zip(NI_No, PN_list, supplier_email, date_list):
        # construct the email item object
        mailItem = olapp.CreateItem(0)
        mailItem.To = email
        mailItem.Subject = 'FAIR rejection notice'
        mailItem.BodyFormat = 1
        mailItem.Body = """
        Hello,
        
        Net-Inspect FAIR #{}, part number {} was reviewed and disapproved on {}.
        Please apply requested amendments per rejection comments highlighted in red and resubmit to Incora for review.'
        
        Thank you
        """.format(item, PN, date.date())


        mailItem.Display()
        mailItem.Save()
        mailItem.Send()
